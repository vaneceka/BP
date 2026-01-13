import re
import zipfile
from pathlib import Path
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils import range_boundaries, column_index_from_string

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


class ExcelDocument:
    def __init__(self, path: str):
        self.path = path
        self.wb = load_workbook(path, data_only=False)
        self.wb_values = load_workbook(path, data_only=True) 

        self.NS = NS
        self._zip = zipfile.ZipFile(path)
        self.workbook_xml = self._load_xml("xl/workbook.xml")
        self.sheets = self._load_sheets_xml()    


    def _load_xml(self, name: str) -> ET.Element:
        with self._zip.open(name) as f:
            return ET.fromstring(f.read())

    def _load_sheets_xml(self) -> dict:
        sheets = {}

        rels = self._load_xml("xl/_rels/workbook.xml.rels")
        rel_map = {
            r.attrib["Id"]: r.attrib["Target"]
            for r in rels.findall("rel:Relationship", self.NS)
        }

        for sheet in self.workbook_xml.findall(".//main:sheet", self.NS):
            name = sheet.attrib["name"]
            r_id = sheet.attrib.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            target = rel_map.get(r_id)
            if not target:
                continue

            xml = self._load_xml(f"xl/{target}")
            sheets[name] = {"xml": xml, "path": target}

        return sheets

    def get_cell_value_cached(self, sheet: str, addr: str):
        return self.wb_values[sheet][addr].value

    def sheet_names(self) -> list[str]:
        return self.wb.sheetnames

    def get_cell(self, address: str, *, include_value=False):
        if "!" not in address:
            raise ValueError("Cell address must be in format 'sheet!A1'")

        sheet, addr = address.split("!", 1)

        if sheet not in self.wb.sheetnames:
            return None

        ws = self.wb[sheet]
        cell = ws[addr]

        if cell.value is None and cell.data_type != "f":
            return None

        data = {
            "sheet": sheet,
            "address": addr,
            "raw_cell": cell,
            "formula": cell.value if cell.data_type == "f" else None,
            "value_cached": self.wb_values[sheet][addr].value,
        }

        if include_value:
            data["value"] = cell.value

        return data

    def save_xml(self, out_dir: str | Path = "debug_excel_xml"):
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        with zipfile.ZipFile(self.path) as z:
            for name in z.namelist():
                if not name.endswith(".xml"):
                    continue

                try:
                    raw = z.read(name)
                    root = ET.fromstring(raw)
                except Exception:
                    continue

                pretty = minidom.parseString(
                    ET.tostring(root, encoding="utf-8")
                ).toprettyxml(indent="  ", encoding="utf-8")

                target = out_dir / name
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(pretty)


#------------
    def get_array_formula_cells(self) -> list[str]:
        cells = []

        for sheet_name, data in self.sheets.items():
            xml = data["xml"]

            for c in xml.findall(".//main:c", self.NS):
                f = c.find("main:f", self.NS)
                if f is not None and f.attrib.get("t") == "array":
                    addr = c.attrib.get("r")
                    cells.append(f"{sheet_name}!{addr}")

        return cells
    
    def get_cell_info(self, sheet: str, addr: str):
        cell = self.get_cell(f"{sheet}!{addr}")
        if cell is None:
            return None

        formula = cell.get("formula")
        value = cell.get("value_cached")

        return {
            "exists": True,
            "formula": formula,
            "value_cached": value,
            "is_error": isinstance(value, str) and value.startswith("#"),
        }
    
    def iter_formulas(self):
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and isinstance(cell.value, str):
                        yield {
                            "sheet": ws.title,
                            "formula": cell.value,
                        }

    def normalize_formula(self, f: str | None) -> str:
        if not f:
            return ""
        return re.sub(r"\s+", "", f).upper()
    
    def get_defined_names(self) -> set[str]:
        return {name.upper() for name in self.wb.defined_names.keys()}

    def cells_with_formulas(self):
        cells = []

        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        cells.append({
                            "sheet": ws.title,
                            "address": cell.coordinate,
                            "formula": cell.value,
                        })

        return cells
        
    def merged_ranges(self, sheet: str):
        ws = self.wb[sheet]
        ranges = []

        for r in ws.merged_cells.ranges:
            ranges.append(range_boundaries(str(r)))

        return ranges
    
    def has_conditional_formatting(self, sheet: str) -> bool:
        if sheet not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet]
        return bool(ws.conditional_formatting._cf_rules)
    
    def check_conditional_formatting(self, sheet: str, expected: list[dict]) -> list[str]:
        ws = self.wb[sheet]
        found = [False] * len(expected)

        for cf, rules in ws.conditional_formatting._cf_rules.items():
            range_str = str(cf.sqref)

            for r in rules:
                if r.type != "cellIs" or not r.operator or not r.formula:
                    continue

                try:
                    value = float(r.formula[0])
                except Exception:
                    continue

                for i, exp in enumerate(expected):
                    if found[i]:
                        continue

                    cell_addr = exp.get("cell")
                    if not cell_addr:
                        continue

                    # CF se musí týkat buňky (nebo alespoň sloupce)
                    if cell_addr not in range_str:
                        col = "".join(ch for ch in cell_addr if ch.isalpha()).upper()
                        if col not in range_str:
                            continue

                    want_op = exp.get("operator")
                    want_val = float(exp.get("value"))

                    # mapování assignment -> openpyxl operator
                    op_map = {
                        "greaterThan": "greaterThan",
                        "lessThan": "lessThan",
                        "greaterThanOrEqual": "greaterThanOrEqual",
                        "lessThanOrEqual": "lessThanOrEqual",
                        "equal": "equal",
                        "notEqual": "notEqual",
                    }

                    if op_map.get(want_op) == r.operator and abs(value - want_val) < 0.01:
                        found[i] = True

        missing = []
        for i, exp in enumerate(expected):
            if not found[i]:
                missing.append(f'{exp["operator"]} {exp["value"]} (XLSX)')

        return missing

    def get_cell_style(self, sheet: str, addr: str) -> dict | None:        
        try:
            ws = self.wb[sheet]
        except KeyError:
            return None

        cell = ws[addr]
        if cell is None:
            return None


        return {
            "number_format": cell.number_format,
            "align_h": cell.alignment.horizontal,
            "bold": bool(cell.font and cell.font.bold),
        }
    
    def iter_cells(self, sheet: str):
        try:
            ws = self.wb[sheet]
        except KeyError:
            return
        for row in ws.iter_rows():
            for cell in row:
                yield cell.coordinate

    def get_cell_value(self, sheet: str, addr: str):
        ws = self.wb[sheet]
        return ws[addr].value
    
    def has_formula(self, sheet: str, addr: str) -> bool:
        ws = self.wb[sheet]
        cell = ws[addr]
        return isinstance(cell.value, str) and cell.value.startswith("=")
    
    def has_chart(self, sheet: str) -> bool:
        try:
            ws = self.wb[sheet]
        except KeyError:
            return False
        return bool(ws._charts)

    def _chart(self, sheet: str):
        ws = self.wb[sheet]
        return ws._charts[0] if ws._charts else None

    def chart_title(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if not c or not c.title or not c.title.tx:
            return None
        return c.title.tx.rich.p[0].r[0].t.strip()

    def chart_x_label(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if c and c.x_axis and c.x_axis.title:
            return c.x_axis.title.tx.rich.p[0].r[0].t
        return None

    def chart_y_label(self, sheet: str) -> str | None:
        c = self._chart(sheet)
        if c and c.y_axis and c.y_axis.title:
            return c.y_axis.title.tx.rich.p[0].r[0].t
        return None

    def chart_has_data_labels(self, sheet: str) -> bool:
        c = self._chart(sheet)
        if not c:
            return False
        return any(s.dLbls and s.dLbls.showVal for s in c.series)
    
    def chart_type(self, sheet: str) -> str | None:
        if sheet not in self.wb.sheetnames:
            return None

        ws = self.wb[sheet]
        if not ws._charts:
            return None

        chart = ws._charts[0]

        # openpyxl typy
        tag = chart.tagname.lower()

        if "bar" in tag:
            return "bar"
        if "line" in tag:
            return "line"
        if "pie" in tag:
            return "pie"

        return tag
    
    def has_3d_chart(self, sheet: str) -> bool:
        if sheet not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet]

        for chart in ws._charts:
            tag = chart.tagname.lower()
            if "3d" in tag:
                return True

        return False
    
    def has_sheet(self, name: str) -> bool:
        return name in self.wb.sheetnames
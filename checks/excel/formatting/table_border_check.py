from checks.base_check import BaseCheck, CheckResult
from openpyxl.utils import range_boundaries

class TableBorderCheck(BaseCheck):
    name = "Chybí vnitřní/vnější ohraničení tabulky"
    penalty = -1
    SHEET = "data"

    # Mapování „zadání → realita openpyxl“
    # Excel UI „tlustá“ často = openpyxl "medium"
    STYLE_EQUIV = {
        "thick": {"thick", "medium"},
        "medium": {"medium"},
        "thin": {"thin"},
        "hair": {"hair"},
        "dotted": {"dotted"},
        "dashDot": {"dashDot"},
        "dashDotDot": {"dashDotDot"},
        "dashed": {"dashed"},
        "double": {"double"},
        "slantDashDot": {"slantDashDot"},
    }

    def _get(self, obj, key):
        return obj.get(key) if isinstance(obj, dict) else getattr(obj, key)

    def _style(self, side):
        return getattr(side, "style", None)

    def _matches(self, actual_style, expected_style):
        allowed = self.STYLE_EQUIV.get(expected_style, {expected_style})
        return actual_style in allowed

    def run(self, document, assignment=None):
        if assignment is None or not hasattr(assignment, "borders"):
            return CheckResult(True, "Chybí definice tabulek – check přeskočen.", 0)

        if self.SHEET not in document.wb.sheetnames:
            return CheckResult(False, f'Chybí list "{self.SHEET}".', self.penalty, fatal=True)

        ws = document.wb[self.SHEET]
        problems = []

        for table in assignment.borders:
            rng = self._get(table, "location")
            outer = self._get(table, "outlineBorderStyle")   # v JSON "thick"
            inner = self._get(table, "insideBorderStyle")    # v JSON "thin"

            min_col, min_row, max_col, max_row = range_boundaries(rng)

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border

                    # outer
                    if r == min_row:
                        if not self._matches(self._style(b.top), outer):
                            problems.append(f"{cell.coordinate}: chybí TOP vnější ({outer})")
                    if r == max_row:
                        if not self._matches(self._style(b.bottom), outer):
                            problems.append(f"{cell.coordinate}: chybí BOTTOM vnější ({outer})")
                    if c == min_col:
                        if not self._matches(self._style(b.left), outer):
                            problems.append(f"{cell.coordinate}: chybí LEFT vnější ({outer})")
                    if c == max_col:
                        if not self._matches(self._style(b.right), outer):
                            problems.append(f"{cell.coordinate}: chybí RIGHT vnější ({outer})")

                    # inner
                    if r > min_row:
                        top_ok = self._matches(self._style(b.top), inner)
                        if not top_ok:
                            up_b = ws.cell(row=r - 1, column=c).border
                            top_ok = self._matches(self._style(up_b.bottom), inner)
                        if not top_ok:
                            problems.append(f"{cell.coordinate}: chybí TOP vnitřní ({inner})")

                    if c > min_col:
                        left_ok = self._matches(self._style(b.left), inner)
                        if not left_ok:
                            left_b = ws.cell(row=r, column=c - 1).border
                            left_ok = self._matches(self._style(left_b.right), inner)
                        if not left_ok:
                            problems.append(f"{cell.coordinate}: chybí LEFT vnitřní ({inner})")

        if problems:
            return CheckResult(
                False,
                "Chybí ohraničení tabulky:\n" + "\n".join("– " + p for p in problems[:15]),
                self.penalty,
                fatal=True
            )

        return CheckResult(True, "Tabulky mají správné vnější i vnitřní ohraničení.", 0)